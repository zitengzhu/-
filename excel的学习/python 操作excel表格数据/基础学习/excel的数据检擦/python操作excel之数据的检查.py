# 导入  读取excel的模块
from openpyxl import load_workbook


# 读取正在使用的excel表格内的sheet表格的信息
def read_active_excel():
    # 1.加载excel数据  括号内的参数是要检查数据的excel表格的名称
    wb = load_workbook('学生信息.xlsx')
    # 2.加载excel完毕后 开始读取表格内正在使用的sheet表格
    sheet = wb.active
    # print(sheet)          # <Worksheet "Sheet">
    # print(type(sheet))    # <class 'openpyxl.worksheet.worksheet.Worksheet'>
    # 发现数据看不懂可以设置循环
    for i in sheet:
        # print(i)
        '''
        (<Cell 'Sheet'.A1>, <Cell 'Sheet'.B1>, <Cell 'Sheet'.C1>, <Cell 'Sheet'.D1>)
        (<Cell 'Sheet'.A2>, <Cell 'Sheet'.B2>, <Cell 'Sheet'.C2>, <Cell 'Sheet'.D2>)
        (<Cell 'Sheet'.A3>, <Cell 'Sheet'.B3>, <Cell 'Sheet'.C3>, <Cell 'Sheet'.D3>)
        '''
        # 发现i的数据是sheet表格内 所有存在数据的单元格的坐标  但每一行依旧是一个小列表
        # 再次嵌套循环 可以获取每一个单元格的坐标
        for i in i:
            # .value  可以拿到单元格内的值
            print(i,i.value)
            '''
            <Cell 'Sheet'.A1> 学生编号
            <Cell 'Sheet'.B1> 姓名
            <Cell 'Sheet'.C1> 年龄
            <Cell 'Sheet'.D1> 性别
            <Cell 'Sheet'.A2> 1
            <Cell 'Sheet'.B2> 张三
            <Cell 'Sheet'.C2> 18
            <Cell 'Sheet'.D2> 男
            <Cell 'Sheet'.A3> 2
            <Cell 'Sheet'.B3> 李四
            <Cell 'Sheet'.C3> 20
            <Cell 'Sheet'.D3> 男
            '''

# 读取所有的sheet表格数据
def read_all_excel():
    # 1.加载要读取的excel表格
    wb = load_workbook('学生信息.xlsx')
    # 2.excel表格加载完毕后开始拿取所有的sheet表格
    sheets = wb.worksheets  # 拿取所有的sheet表格
    # 设置循环拿取数值
    for data in sheets:
        # print(i) <Worksheet "Sheet">   第一个循环拿到所有sheet表格的表名
        for i in data:
            # print(i)  第二个循环拿到的是 每一行的数据
            '''
            (<Cell 'Sheet'.A1>, <Cell 'Sheet'.B1>, <Cell 'Sheet'.C1>, <Cell 'Sheet'.D1>)
            (<Cell 'Sheet'.A2>, <Cell 'Sheet'.B2>, <Cell 'Sheet'.C2>, <Cell 'Sheet'.D2>)
            (<Cell 'Sheet'.A3>, <Cell 'Sheet'.B3>, <Cell 'Sheet'.C3>, <Cell 'Sheet'.D3>)
            '''
            for obj in i:
                print(obj)  # 第三个循环得到的是每一个单元格的数据   如果想要拿到值那就是加上.value
                '''
                <Cell 'Sheet'.A1>
                <Cell 'Sheet'.B1>
                <Cell 'Sheet'.C1>
                <Cell 'Sheet'.D1>
                <Cell 'Sheet'.A2>
                <Cell 'Sheet'.B2>
                <Cell 'Sheet'.C2>
                <Cell 'Sheet'.D2>
                <Cell 'Sheet'.A3>
                <Cell 'Sheet'.B3>
                <Cell 'Sheet'.C3>
                <Cell 'Sheet'.D3>
                '''

def read_specify_excel():
    # 加载excel表格的数据 括号内的参数就是excel表格的名称
    wb = load_workbook('学生信息.xlsx')
    # 加载表格完毕之后开始读取 指定sheet表格的数据
    sheet = wb['Sheet']  # 括号内就是想要读取的sheet表格的内容
    # 设置完毕后开始设置循环获取数据
    for i in sheet:
        # print(i) 循环获取的是每一行的数据
        for i in i:
            print(i)  # 循环获取的是每一个单元格的数据





