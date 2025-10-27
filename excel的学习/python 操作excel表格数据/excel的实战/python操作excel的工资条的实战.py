'''
要想打印每一个人的工资条
1.获取excel表格中每个人的数据需要设置循环
2.设置循环获取到数据后需要建立新的sheet表格来存储每个人的数据
'''
import time

from openpyxl import Workbook
from openpyxl import load_workbook
import os


def read_excel():
    # 加载excel表格  因为是要从表格中获取数据所以不能用Workbook
    web = load_workbook('工资明细.xlsx')  # 获取 工资明细.xlsx 内的数据
    # 加载表格完成后开始获取指定sheet表格的数据
    sheet = web['Sheet']
    # 数据获取完毕之后开始设置循环来获取数据
    # 前三行数据不是目前的目标数据先进行过滤
    count = 0
    for i in sheet:
        count = count+1
        if count < 4:
            continue
        else:
            # print(i) # 打印后发现获取到了每一行的数据 要定义一个列表去接受数据
            # print('-----------')
            data_list = []
            for j in i:
                # print(j) # 获取到了每一个单元格数据
                # 获取到数据后添加进入表格
                data_list.append(j.value)
            # 获取到了每一个人的数据开始创建列表来保存每一个人的数据
            write_excel(data_list)


def write_excel(data_list):  # 括号内进行传参可以让 write_excel 函数有 data_list 的数据
    # 1.创建一个excel表格
    web = Workbook()
    # 2.创建完毕之后开始创建sheet表格
    sheet = web.active
    # 表格创建完之后开始导入第三行过滤掉的表头信息  数据的形式一定是列表
    title = ["序号", "姓名", "岗位工资", "工龄冿贴", "学历职称冿贴", "合计", "养老保险", "住房公积金","失业保险", "工会会员年费", "处分人员罚款", "实发工资"]
    # 将设置好的表头添加到sheet表格内
    sheet.append(title)
    # 将第一个函数处理好的数据放进sheet表格内
    sheet.append(data_list)
    # 要建立sheet列表存放每一个人的数据 所以到获取到每一个人的名称 方便标记sheet表格
    name = data_list[1]
    # sheet表格导入完毕开始保存文件
    if not os.path.exists('money'):
        os.mkdir('money')
    # 设置了文件夹开始保存sheet表格
    web.save(f'money/{name}.xlsx')
    print(f'{name}'+'工资条制作完毕-----------------')
    time.sleep(1)
read_excel()